import os
from openpyxl import Workbook

def classificar_arquivo(nome_arquivo):
    if nome_arquivo.startswith('BX'):
        return 'Boxe'
    elif nome_arquivo.startswith('CA'):
        return 'Capoeira'
    else:
        return 'Outros'

pasta = r'Pasta a ser escaneada'

planilha = Workbook()
planilha_ativa = planilha.active

arquivos = os.listdir(pasta)

for indice, arquivo in enumerate(arquivos, start=1):
    categoria = classificar_arquivo(arquivo)
    planilha_ativa.cell(row=indice, column=1).value = arquivo
    planilha_ativa.cell(row=indice, column=2).value = categoria

planilha.save(r'Pasta onde vai ser salvo o arquivo/aqui_o_nome_do_excel.xlsx')
